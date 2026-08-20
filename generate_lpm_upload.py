"""
generate_lpm_upload.py

Reads the "Tracking Table" sheet from a Goal Builder .xlsm file and generates
LPM-format CSV files ready for upload to LiquidPerform.

Usage:
    python generate_lpm_upload.py <goal_builder.xlsm> [output_csv] [collection_report.xlsx]

One CSV is generated per tracker group. Trackers are grouped by:
    Goal Group + SPP Tier + Objective Type + Program Period (Start/End) + Unsold Period

State abbreviation is derived from the Goal Builder filename (e.g. "SD SPP..." -> "SD").
Collection IDs are looked up from the Collection ID List sheet in the collection report.
"""

import csv
import os
import sys
from collections import defaultdict
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required.  Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Site-wide defaults
# ---------------------------------------------------------------------------
CONFIG = {
    "basis_flag":                     "TRUE",
    "score_by_td_linx_customer_code": "",
    "attainment_org_level":           "Salesperson",
    "recalculate_until_date":         "365",
    "exclude_from_os_sales_reports":  "FALSE",
    "salesforce_collection_ids":      "",
    "send_to_proof":                  "SendStartingInActive",
    "send_to_proof_date":             "",
    "distribution_level_path":        "Salesperson",
}

# Tracking Table "Objective Type" (col 4) -> LPM goal_type
OBJECTIVE_TYPE_MAP = {
    "Volume (Cases)":   "VolumeCases",
    "New POD":          "DistributionNewPODs",
    "POD":              "DistributionPODs",
    "New ACS":          "DistributionNewACS",
    "ACS":              "DistributionACS",
    "Revenue":          "VolumeRevenue",
    "DREV":             "VolumeRevenueDREV",
    "GP":               "VolumeGP",
    "Volume Combos":    "VolumeCombos",
    "Volume Points":    "VolumePoints",
    "Reorder":          "DistributionReorder",
}

# Tracker type short labels used in output filenames
TRACKER_TYPE_LABELS = {
    "VolumeCases":          "Vol",
    "DistributionNewPODs":  "NPOD",
    "DistributionPODs":     "POD",
    "DistributionNewACS":   "NACS",
    "DistributionACS":      "ACS",
    "VolumeCombos":         "Combo",
    "VolumePoints":         "Pts",
    "DistributionReorder":  "Reorder",
}

DISTRIBUTION_TYPES = {
    "DistributionNewPODs", "DistributionPODs",
    "DistributionACS", "DistributionNewACS",
    "DistributionReorder",
}

VOLUME_TYPES = {
    "VolumeCases", "VolumeRevenue", "VolumeRevenueDREV",
    "VolumeGP", "VolumeCombos", "VolumePoints",
}

# Only NPOD and NACS support unsold periods
UNSOLD_TYPES = {"DistributionNewPODs", "DistributionNewACS"}
# Only NPOD and POD support POD attributes
POD_ATTR_TYPES = {"DistributionNewPODs", "DistributionPODs"}

# Tracking Table "Goal Distribution" (col 20) value that marks a fixed-per-salesperson goal
FIXED_GOAL_DISTRIBUTION = "Fixed Goal per Salesperson"
# ...and the value that marks a goal split evenly among reps. Both of these use the
# Market Segment Goal itself as the per-rep target rather than "Min Goal per Rep".
EVEN_GOAL_DISTRIBUTION = "Even Goal (Unit Goal split evenly among reps)"
NO_MIN_FALLBACK_DISTRIBUTIONS = {FIXED_GOAL_DISTRIBUTION, EVEN_GOAL_DISTRIBUTION}

# Tracking Table "Measure" (col 10) -> LPM goal_uom
UOM_MAP = {
    "9L":      "NineLiter",
    "Cases":   "Cases",
    "Dec":     "Cases",
    "dcml":    "Cases",
    "QTY_DEC": "Cases",
    "STD":          "STD",
    "Standard Cs":  "STD",
    "STD_CASE":     "STD",
    "BATF":    "BATF",
    "Bottles":        "Bottles",
    "SUPP_CASE":      "SupplierRatio",
    "Supplier Ratio": "SupplierRatio",
    "Supp":           "SupplierRatio",
}

# SPP Tier (col 2) -> program_class
PROGRAM_CLASS_MAP = {
    "Anchor": "ProgramClass.Site.Sppa",
    "Flex":   "ProgramClass.Site.Sppf",
}

# Objective types and measures to skip entirely
SKIP_OBJECTIVE_TYPES = {
    "SPP My Sales",
    "Proof Comm Invoice Freq",
    "Proof Com Activation",
    "Proof Comm Proposals",
    "SET Backbar",
    "SET Menu",
    "SET - Use Expanded Columns",
}
SKIP_MEASURES        = {"DigComRev"}

# LPM CSV column order (must match upload spec)
CSV_COLUMNS = [
    "goal_category", "goal_name", "tpm_nav_ref", "goal_type", "goal_description",
    "goal_start_date", "goal_end_date", "basis_flag", "score_by_td_linx_customer_code",
    "attainment_org_level", "goal_uom", "recalculate_until_date",
    "exclude_from_os_sales_reports", "salesforce_collection_ids", "program_class",
    "send_to_proof", "send_to_proof_date", "unsold_start_date", "unsold_end_date",
    "product_collection_id", "customer_collection_id", "distribution_target",
    "min_objective_target", "distribution_level_path", "pod_attribute", "achievement_min",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_yyyymm(value):
    """Convert a cell value (datetime, int YYYYMM, or string) to a YYYYMM string."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m")
    s = str(value).strip()
    if not s:
        return ""
    if len(s) == 6 and s.isdigit():
        return s
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y%m")
        except ValueError:
            pass
    return s


def next_future_month_yyyymm():
    """Return the first month that is strictly in the future from today (YYYYMM)."""
    today = date.today()
    # Next month after today
    if today.month == 12:
        return f"{today.year + 1}01"
    return f"{today.year}{today.month + 1:02d}"


def _yyyymm_subtract_months(yyyymm, months):
    """Return YYYYMM shifted back by the given number of whole months."""
    year  = int(yyyymm[:4])
    month = int(yyyymm[4:6])
    month -= months
    while month <= 0:
        month += 12
        year  -= 1
    return f"{year}{month:02d}"


def compute_unsold_dates(unsold_prd, goal_start_yyyymm, supplier_name="", goal_end_yyyymm=""):
    """
    Compute (unsold_start_yyyymm, unsold_end_yyyymm) from the unsold period value
    and the tracker's incentive start/end months.

    Rules:
      - R12: unsold_end = one month prior to tracking end; unsold_start = 12 months prior to that.
             e.g. July-July tracker (end=202707) -> unsold_end=202706, unsold_start=202507.
      - CYTD: January of the incentive start year through one month prior to start.
      - FYTD: Supplier's fiscal start month (from FISCAL_LOOKUP) through one month
              prior to incentive start. Falls back to CYTD if supplier not found.
      - Numeric day value (e.g. "30", "90", "30 days"): assume 30 days = 1 month.
        unsold_end   = one month prior to incentive start.
        unsold_start = N months prior to incentive start.
        e.g. 30 days -> start=202607 -> 202606 to 202606
             90 days -> start=202607 -> 202604 to 202606
      - Anything else: return ("", "").
    """
    if not unsold_prd or not goal_start_yyyymm or len(goal_start_yyyymm) < 6:
        return "", ""

    prd = unsold_prd.strip()

    if prd.upper() == "R12":
        unsold_end = _yyyymm_subtract_months(goal_start_yyyymm, 1)
        unsold_start = _yyyymm_subtract_months(unsold_end, 11)
        return unsold_start, unsold_end

    unsold_end = _yyyymm_subtract_months(goal_start_yyyymm, 1)

    if prd.upper() == "CYTD":
        year = goal_start_yyyymm[:4]
        return f"{year}01", unsold_end

    if prd.upper() == "FYTD":
        fiscal_start = FISCAL_LOOKUP.get(supplier_name.lower(), "")
        if fiscal_start:
            return fiscal_start, unsold_end
        # Fall back to CYTD if supplier not in lookup
        year = goal_start_yyyymm[:4]
        return f"{year}01", unsold_end

    # Extract leading number — handles "30", "90 days", "90 Days", "30 day", etc.
    import re as _re
    m = _re.match(r"(\d+)", prd)
    if m:
        days = int(m.group(1))
        months = max(1, round(days / 30))
        unsold_start = _yyyymm_subtract_months(goal_start_yyyymm, months)
        return unsold_start, unsold_end

    return "", ""


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def _numeric_str(val):
    """Return the numeric portion of a value. Extracts leading number from strings like '5 per SC'."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    try:
        float(s)
        return s
    except ValueError:
        import re
        m = re.match(r"(-?\d+\.?\d*)", s)
        return m.group(1) if m else ""


VALID_DIST_GOAL_FROM = {
    "Org", "Corporate", "Region", "State", "Site",
    "Market", "Section", "Territory", "Team", "Salesperson",
}

VALID_POD_ATTRIBUTES = {
    "PhSubGroup", "VarietalSize", "Flavor", "PhSuperGroup", "PhGroup",
    "PodId", "PodName", "ProductId", "ProductSize", "SubgroupSize",
    "FlavorSize", "ItemDescVintageRoll", "Varietal",
}

POD_ATTRIBUTE_MAP = {
    "sub group":                              "PhSubGroup",
    "varietal size":                          "VarietalSize",
    "flavor":                                 "Flavor",
    "super group":                            "PhSuperGroup",
    "group":                                  "PhGroup",
    "supplier pod id":                        "PodId",
    "item":                                   "ProductId",
    "size":                                   "ProductSize",
    "sub group size":                         "SubgroupSize",
    "flavor size":                            "FlavorSize",
    "item roll size (vintage trim)/pim item": "ItemDescVintageRoll",
    "varietal":                               "Varietal",
}

def normalize_pod_attribute(val):
    s = safe_str(val)
    if not s or s == "-":
        return ""
    # Apply known aliases first
    mapped = POD_ATTRIBUTE_MAP.get(s.lower())
    if mapped:
        return mapped
    return s


def ptg_name_from_row(ptg_name_col, selection_col, supplier_col):
    """
    PTG name: col 12 preferred; fall back to col 15 (Selection), then col 14 (Supplier).
    Truncate to 256 chars (LPM limit).
    """
    for v in (ptg_name_col, selection_col, supplier_col):
        s = safe_str(v)
        if s:
            return s[:256]
    return ""


def safe_filename(s):
    for ch in r'\/:*?"<>| ':
        s = s.replace(ch, "_")
    return s


# ---------------------------------------------------------------------------
# Collection ID lookup
# ---------------------------------------------------------------------------

def derive_state_from_filename(filepath):
    """Extract state abbreviation from the Goal Builder filename e.g. 'SD SPP...' -> 'SD'."""
    name = os.path.splitext(os.path.basename(filepath))[0]
    # First token before a space is the state abbreviation
    return name.split()[0].upper()


def _parse_collection_sheet_raw(file_source, state):
    """
    Parse the 'Collection ID List' sheet by reading the raw XML inside the xlsx zip.
    This avoids openpyxl's float conversion entirely, preserving all 16 digits of
    collection IDs.

    Returns {lowercase_name: id_string}
    """
    import zipfile
    import xml.etree.ElementTree as ET
    import io as _io

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    # Accept a file path or bytes
    if isinstance(file_source, (bytes, bytearray)):
        zf = zipfile.ZipFile(_io.BytesIO(file_source))
    else:
        zf = zipfile.ZipFile(file_source)

    # Map sheet name -> actual file path via workbook.xml + workbook.xml.rels
    with zf.open("xl/workbook.xml") as f:
        wb_tree = ET.parse(f)

    # rId -> sheet name
    rid_to_name = {}
    REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    for sheet in wb_tree.findall(f".//{{{NS}}}sheet"):
        rid = sheet.get(f"{{{REL_NS}}}id")
        rid_to_name[rid] = sheet.get("name")

    # rId -> actual file path from relationships file
    rid_to_path = {}
    rels_path = "xl/_rels/workbook.xml.rels"
    if rels_path in zf.namelist():
        with zf.open(rels_path) as f:
            rels_tree = ET.parse(f)
        RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
        for rel in rels_tree.findall(f"{{{RELS_NS}}}Relationship"):
            target = rel.get("Target")
            path = target.lstrip("/") if target.startswith("/") else "xl/" + target
            rid_to_path[rel.get("Id")] = path

    target_name = "Collection ID List"
    candidate = None
    for rid, name in rid_to_name.items():
        if name == target_name:
            candidate = rid_to_path.get(rid)
            break

    if not candidate or candidate not in zf.namelist():
        zf.close()
        return {}

    # Read shared strings table
    shared_strings = []
    if "xl/sharedStrings.xml" in zf.namelist():
        with zf.open("xl/sharedStrings.xml") as f:
            ss_tree = ET.parse(f)
        for si in ss_tree.findall(f"{{{NS}}}si"):
            parts = si.findall(f".//{{{NS}}}t")
            shared_strings.append("".join(p.text or "" for p in parts))

    def cell_value(c_elem):
        t = c_elem.get("t", "n")
        if t == "inlineStr":
            is_elem = c_elem.find(f"{{{NS}}}is")
            if is_elem is None:
                return ""
            parts = is_elem.findall(f".//{{{NS}}}t")
            return "".join(p.text or "" for p in parts)
        v_elem = c_elem.find(f"{{{NS}}}v")
        if v_elem is None or v_elem.text is None:
            return ""
        if t == "s":
            idx = int(v_elem.text)
            return shared_strings[idx] if idx < len(shared_strings) else ""
        return v_elem.text.strip()

    lookup = {}

    with zf.open(candidate) as f:
        ws_tree = ET.parse(f)

    rows = ws_tree.findall(f".//{{{NS}}}row")
    for row_elem in rows:
        r_num = int(row_elem.get("r", 0))
        if r_num < 2:
            continue
        cells = row_elem.findall(f"{{{NS}}}c")
        # Map column letter -> value
        col_vals = {}
        for c in cells:
            ref = c.get("r", "")
            col_letter = "".join(ch for ch in ref if ch.isalpha()).upper()
            col_vals[col_letter] = cell_value(c)

        row_state = col_vals.get("A", "").strip().upper()
        name      = col_vals.get("B", "").strip()
        coll_id   = col_vals.get("C", "").strip()

        if row_state != state.upper():
            continue
        if not name or not coll_id:
            continue
        if "(do not use)" in name.lower():
            continue

        # Strip trailing .0 if the ID came through as "4844571982686355.0"
        if coll_id.endswith(".0"):
            coll_id = coll_id[:-2]

        lookup[name.lower()] = coll_id

    zf.close()
    return lookup


def load_collection_lookup(collection_path, state):
    """
    Build a dict mapping lowercase collection name -> collection ID for the given state.
    Reads raw XML from the xlsx zip to avoid openpyxl float64 precision loss.
    """
    if not collection_path or not os.path.isfile(collection_path):
        return {}
    return _parse_collection_sheet_raw(collection_path, state)
    return lookup


def load_fiscal_lookup(fiscal_path):
    """
    Build a dict mapping lowercase supplier name -> fiscal start YYYYMM.
    Reads 'Supplier Fiscal Start Month.xlsx' — header on row 3,
    data from row 4: col B = supplier number, col C = supplier name, col D = fiscal start YYYYMM.
    """
    if not fiscal_path or not os.path.isfile(fiscal_path):
        return {}
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(fiscal_path, data_only=True)
    ws = wb.active
    lookup = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        name  = str(row[2]).strip() if row[2] else ""   # col C
        start = str(row[3]).strip() if row[3] else ""   # col D
        if not name or not start or name == "None":
            continue
        # Normalize YYYYMM — may come back as float e.g. 202601.0
        if start.endswith(".0"):
            start = start[:-2]
        if len(start) == 6 and start.isdigit():
            lookup[name.lower()] = start
    return lookup


# ---------------------------------------------------------------------------
# Load Tracking Table
# ---------------------------------------------------------------------------

def load_tracking_table(wb_path):
    """
    Read the Tracking Table sheet and return (records, skipped_rows).
    Each record is a dict of parsed/mapped fields ready for CSV output.
    """
    # Load twice: formula values for text/dropdown cells (avoids stale cached values),
    # data_only for date cells which need calculated results
    wb_formula = openpyxl.load_workbook(wb_path, keep_vba=True, data_only=False)
    wb = openpyxl.load_workbook(wb_path, keep_vba=True, data_only=True)
    ws_formula = wb_formula["Tracking Table"]

    if "Tracking Table" not in wb.sheetnames:
        raise ValueError(
            f"No 'Tracking Table' sheet found in {wb_path}.\n"
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb["Tracking Table"]
    # Capture header row for the skipped-rows export
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    records = []
    skipped = []  # each entry: {"row_num", "reason", "raw_row"}

    def make_skip(row_num, reason, raw):
        return {"row_num": row_num, "reason": reason, "raw_row": raw}

    for r in range(2, ws.max_row + 1):
        def cv(c):
            # Use formula workbook for text/dropdown columns to avoid stale cached values
            # Use data_only workbook for date columns (cols 5-8) which need calculated results
            if c in (5, 6, 7, 8):
                return ws.cell(r, c).value
            v = ws_formula.cell(r, c).value
            # If formula workbook returns a formula string, fall back to cached value
            if isinstance(v, str) and v.startswith("="):
                return ws.cell(r, c).value
            return v

        goal_group  = safe_str(cv(1))
        spp_tier    = safe_str(cv(2))
        goal_bucket = safe_str(cv(3))
        obj_type    = safe_str(cv(4)).rstrip(" -")
        start_raw   = cv(5)
        end_raw          = cv(6)
        basis_start_raw  = cv(7)
        basis_end_raw    = cv(8)
        unsold_prd  = safe_str(cv(9))
        measure     = safe_str(cv(10))
        mkt_seg     = cv(11)
        # Excel stores percentage goals as decimals (7.44% -> 0.0744) but reports
        # the format as General when styles are applied via named styles — openpyxl
        # can't detect it. Heuristic: if value is a non-integer between -1 and 1
        # (exclusive), treat it as a percentage and multiply by 100.
        if isinstance(mkt_seg, str) and mkt_seg.strip().upper() == "FLAT":
            mkt_seg = 0
        elif isinstance(mkt_seg, float) and -1 < mkt_seg < 1 and mkt_seg != 0:
            mkt_seg = round(mkt_seg * 100, 4)
        ptg_name_v  = cv(12)
        level_detail = safe_str(cv(13))
        supplier    = cv(14)
        selection       = cv(15)
        # cols 16-19 not used
        goal_distribution = safe_str(cv(20))
        qualifier       = cv(21)  # Qualifier -> achievement_min (numeric only)
        min_goal_per_rep = cv(22) # Min Goal per Rep -> min_objective_target
        # col 23 = Min Cases, col 24 = Min Facings (not used)
        pod_attr        = cv(25)
        # col 26 not used

        raw_row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

        # Skip completely blank rows (no logging)
        if not goal_group and ptg_name_v is None and selection is None:
            continue

        # Skip Digital goals (no logging)
        if measure in SKIP_MEASURES:
            continue

        # Skip "Select:" bucket rows with no objective type (no logging)
        if goal_bucket == "Select:" and not obj_type:
            continue

        # Skip known objective types — log for prompt callout
        if obj_type in SKIP_OBJECTIVE_TYPES:
            skipped.append(make_skip(r, f"skipped objective type: {obj_type!r}", raw_row))
            continue

        # Map objective type
        goal_type = OBJECTIVE_TYPE_MAP.get(obj_type)
        if not goal_type:
            skipped.append(make_skip(r, f"unknown objective type: {obj_type!r}", raw_row))
            continue

        # Coerce goal_type based on unsold period:
        #   POD/ACS + non-None unsold  -> NPOD/NACS
        #   NPOD/NACS + "None" unsold  -> POD/ACS
        _unsold_present = unsold_prd and unsold_prd.strip().upper() != "NONE"
        _unsold_none    = unsold_prd and unsold_prd.strip().upper() == "NONE"
        if goal_type == "DistributionPODs" and _unsold_present:
            goal_type = "DistributionNewPODs"
        elif goal_type == "DistributionACS" and _unsold_present:
            goal_type = "DistributionNewACS"
        elif goal_type == "DistributionNewPODs" and _unsold_none:
            goal_type = "DistributionPODs"
        elif goal_type == "DistributionNewACS" and _unsold_none:
            goal_type = "DistributionACS"

        # Resolve PTG name — explicit name always wins; fall back to "Total <type>" only if blank
        name = ptg_name_from_row(ptg_name_v, selection, supplier)
        if not name and level_detail == "Total":
            name = f"Total {obj_type}"
        if not name:
            skipped.append(make_skip(r, "no PTG name, selection, or supplier", raw_row))
            continue

        # Map UOM; revenue/GP types always blank; distribution types default to Cases
        if goal_type in {"VolumeRevenue", "VolumeRevenueDREV", "VolumeGP", "VolumeCombos", "VolumePoints"}:
            goal_uom = ""
        else:
            goal_uom = UOM_MAP.get(measure, "")
            if not goal_uom and goal_type in DISTRIBUTION_TYPES:
                goal_uom = "Cases"

        # Dates
        start_yyyymm       = to_yyyymm(start_raw)
        end_yyyymm         = next_future_month_yyyymm()
        basis_start_yyyymm = to_yyyymm(basis_start_raw)
        basis_end_yyyymm   = to_yyyymm(basis_end_raw)

        # Validate POD attribute
        pod_attr_clean = normalize_pod_attribute(pod_attr)
        if pod_attr_clean and pod_attr_clean not in VALID_POD_ATTRIBUTES:
            skipped.append(make_skip(r, f"invalid POD attribute: {pod_attr_clean!r}", raw_row))
            continue

        records.append({
            "row_num":       r,
            "goal_group":    goal_group,
            "spp_tier":      spp_tier,
            "goal_type":     goal_type,
            "goal_uom":      goal_uom,
            "program_class": PROGRAM_CLASS_MAP.get(spp_tier, ""),
            "start_yyyymm":        start_yyyymm,
            "end_yyyymm":          end_yyyymm,
            "basis_start_yyyymm":  basis_start_yyyymm,
            "basis_end_yyyymm":    basis_end_yyyymm,
            "unsold_prd":          unsold_prd,
            "supplier":      safe_str(supplier),
            "ptg_name":      name,
            "mkt_seg_goal":       _numeric_str(mkt_seg),
            "pod_attribute":      pod_attr_clean,
            "min_goal_per_rep":   _numeric_str(min_goal_per_rep),
            "qualifier":          _numeric_str(qualifier),
            "goal_distribution":  goal_distribution,
        })

    return records, skipped, header_row


# ---------------------------------------------------------------------------
# Group records
# ---------------------------------------------------------------------------

def group_key(rec):
    # Unsold period only applies to NPOD/NACS — don't let it split other trackers
    unsold = rec["unsold_prd"] if rec["goal_type"] in UNSOLD_TYPES else ""
    return (
        rec["goal_group"],
        rec["spp_tier"],
        rec["goal_type"],
        rec["goal_uom"],
        rec["start_yyyymm"],
        rec["end_yyyymm"],
        rec["basis_start_yyyymm"],
        rec["basis_end_yyyymm"],
        unsold,
    )


def group_records(records):
    groups = defaultdict(list)
    order  = []
    for rec in records:
        key = group_key(rec)
        if key not in groups:
            order.append(key)
        groups[key].append(rec)
    return order, groups


# ---------------------------------------------------------------------------
# Build CSV rows
# ---------------------------------------------------------------------------

COLLECTION_LOOKUP = {}  # populated in main() after loading the collection file
FISCAL_LOOKUP     = {}  # supplier name (lowercase) -> fiscal start YYYYMM


def build_tracker_row(key, recs):
    goal_group, spp_tier, goal_type, goal_uom_key, start, end, basis_start, basis_end, unsold_prd = key

    # UOM: use most common value across the group's records
    uom_counts = defaultdict(int)
    for r in recs:
        if r["goal_uom"]:
            uom_counts[r["goal_uom"]] += 1
    goal_uom = max(uom_counts, key=uom_counts.get) if uom_counts else ""

    program_class = recs[0]["program_class"]

    # Unsold dates — NPOD and NACS only
    unsold_start, unsold_end = "", ""
    if unsold_prd and goal_type in UNSOLD_TYPES:
        supplier_name = recs[0].get("supplier", "")
        unsold_start, unsold_end = compute_unsold_dates(unsold_prd, start, supplier_name, end)

    return {
        "goal_category":                  "Tracker",
        "goal_name":                      goal_group,
        "tpm_nav_ref":                    "",
        "goal_type":                      goal_type,
        "goal_description":               "",
        "goal_start_date":                start,
        "goal_end_date":                  end,
        "basis_flag":                     CONFIG["basis_flag"],
        "score_by_td_linx_customer_code": CONFIG["score_by_td_linx_customer_code"],
        "attainment_org_level":           CONFIG["attainment_org_level"],
        "goal_uom":                       goal_uom,
        "recalculate_until_date":         CONFIG["recalculate_until_date"],
        "exclude_from_os_sales_reports":  CONFIG["exclude_from_os_sales_reports"],
        "salesforce_collection_ids":      COLLECTION_LOOKUP.get(goal_group.lower(), CONFIG["salesforce_collection_ids"]),
        "program_class":                  program_class,
        "send_to_proof":                  CONFIG["send_to_proof"],
        "send_to_proof_date":             CONFIG["send_to_proof_date"],
        "unsold_start_date":              unsold_start,
        "unsold_end_date":                unsold_end,
        "product_collection_id":          "",
        "customer_collection_id":         "",
        "distribution_target":            "",
        "min_objective_target":           "",
        "distribution_level_path":        "",
        "pod_attribute":                  "",
        "achievement_min":                "",
    }


def build_ptg_row(rec):
    is_dist   = rec["goal_type"] in POD_ATTR_TYPES
    is_volume = rec["goal_type"] in VOLUME_TYPES
    is_fixed_or_even = rec["goal_distribution"] in NO_MIN_FALLBACK_DISTRIBUTIONS
    distribution_target = rec["mkt_seg_goal"]
    min_objective_target = distribution_target if is_fixed_or_even else (rec["min_goal_per_rep"] or "1")
    return {
        "goal_category":                  "PTG",
        "goal_name":                      rec["ptg_name"],
        "tpm_nav_ref":                    "",
        "goal_type":                      "",
        "goal_description":               "",
        "goal_start_date":                "",
        "goal_end_date":                  "",
        "basis_flag":                     "",
        "score_by_td_linx_customer_code": "",
        "attainment_org_level":           "",
        "goal_uom":                       "",
        "recalculate_until_date":         "",
        "exclude_from_os_sales_reports":  "",
        "salesforce_collection_ids":      "",
        "program_class":                  "",
        "send_to_proof":                  "",
        "send_to_proof_date":             "",
        "unsold_start_date":              "",
        "unsold_end_date":                "",
        "product_collection_id":          "",
        "customer_collection_id":         "",
        "distribution_target":            distribution_target,
        "min_objective_target":           min_objective_target,
        "distribution_level_path":        CONFIG["distribution_level_path"],
        "pod_attribute":                  (rec["pod_attribute"] or "ProductId") if is_dist else "",
        "achievement_min":                "" if is_volume else rec["qualifier"],
    }


# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------

def generate_output(order, groups, output_path):
    total_trackers = 0
    total_ptgs = 0

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, quoting=csv.QUOTE_ALL)
        writer.writeheader()

        for key in order:
            recs = groups[key]
            tracker_row = build_tracker_row(key, recs)
            ptg_rows    = [build_ptg_row(r) for r in recs]
            writer.writerow(tracker_row)
            for ptg in ptg_rows:
                writer.writerow(ptg)
            total_trackers += 1
            total_ptgs += len(ptg_rows)

    return total_trackers, total_ptgs


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def write_skipped_csv(skipped_path, skipped, header_row):
    """Write skipped rows to a CSV with an extra 'skip_reason' column."""
    with open(skipped_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(["skip_reason"] + header_row)
        for entry in skipped:
            writer.writerow([entry["reason"]] + entry["raw_row"])


def print_summary(output_path, skipped_path, total_trackers, total_ptgs, skipped):
    print()
    print("=" * 65)
    print("LPM Upload Generation Summary")
    print("=" * 65)
    print(f"Output file : {output_path}")
    print(f"Trackers    : {total_trackers}")
    print(f"PTGs        : {total_ptgs}")
    if skipped:
        print(f"\n*** {len(skipped)} row(s) were skipped and need review ***")
        print(f"Skipped rows saved to: {skipped_path}")
        print()
        print(f"  {'Row':<5}  {'Goal Group':<20}  Reason")
        print(f"  {'-'*5}  {'-'*20}  {'-'*40}")
        for entry in skipped:
            print(f"  {entry['row_num']:<5}  {entry['raw_row'][0] or '':<20}  {entry['reason']}")
    else:
        print("\nAll rows processed successfully — no skipped rows.")
    print("=" * 65)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(
            "Usage: python generate_lpm_upload.py <goal_builder.xlsm> [output.csv] [collection_report.xlsx]",
            file=sys.stderr,
        )
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.isfile(input_path):
        print(f"ERROR: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Output path: explicit arg, or same folder as input with .csv extension
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base = os.path.splitext(os.path.abspath(input_path))[0]
        output_path = base + "_lpm_upload.csv"

    # Optional collection report
    collection_path = sys.argv[3] if len(sys.argv) >= 4 else None

    # Auto-detect collection report in same folder if not provided
    if not collection_path:
        folder = os.path.dirname(os.path.abspath(input_path))
        import glob as _glob
        candidates = _glob.glob(os.path.join(folder, "*Collection*Report*.xlsx"))
        if candidates:
            collection_path = candidates[0]

    base = os.path.splitext(output_path)[0]
    skipped_path = base + "_skipped.csv"

    # Load fiscal start month lookup
    global FISCAL_LOOKUP
    fiscal_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Supplier Fiscal Start Month.xlsx")
    if not os.path.isfile(fiscal_path):
        # Fall back to same folder as input file
        fiscal_path = os.path.join(os.path.dirname(os.path.abspath(input_path)), "Supplier Fiscal Start Month.xlsx")
    if os.path.isfile(fiscal_path):
        FISCAL_LOOKUP = load_fiscal_lookup(fiscal_path)
        print(f"Fiscal lookup loaded: {len(FISCAL_LOOKUP)} suppliers  ({os.path.basename(fiscal_path)})")
    else:
        print("Fiscal lookup not found — FYTD unsold periods will fall back to CYTD.")

    # Load collection ID lookup
    global COLLECTION_LOOKUP
    state = derive_state_from_filename(input_path)
    if collection_path and os.path.isfile(collection_path):
        COLLECTION_LOOKUP = load_collection_lookup(collection_path, state)
        print(f"State: {state}  |  Collection IDs loaded: {len(COLLECTION_LOOKUP)}  ({os.path.basename(collection_path)})")
        for name, cid in sorted(COLLECTION_LOOKUP.items()):
            print(f"  {name} -> {cid}")
    else:
        print(f"State: {state}  |  No collection report found — salesforce_collection_ids will be blank.")

    print(f"Reading: {input_path}")
    records, skipped, header_row = load_tracking_table(input_path)

    if not records:
        print("ERROR: No valid records found in Tracking Table after filtering.", file=sys.stderr)
        sys.exit(1)

    order, groups = group_records(records)
    total_trackers, total_ptgs = generate_output(order, groups, output_path)

    if skipped:
        write_skipped_csv(skipped_path, skipped, header_row)

    print_summary(output_path, skipped_path, total_trackers, total_ptgs, skipped)


if __name__ == "__main__":
    main()
