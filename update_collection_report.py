"""
update_collection_report.py

Rebuilds "LPM Salesforce and Overlay Collection Report.xlsx" (the collection
report bundled with the SPP LPM Upload Generator) from a fresh Salesforce
export.

Usage:
    python update_collection_report.py <rep_list_export.xlsx> [overlay_export.xlsx]

If the rep list export contains both sheets (LPM Salesforce Collection Rep
List + Overlay Collection Report), the second argument is not needed.

Rebuilds three sheets in the bundled report:
    - "LPM Salesforce Collection Rep L"  <- raw rep list rows
    - "Overlay Report"                   <- raw overlay rows
    - "Collection ID List"               <- unique (state, name, id) derived
                                             from the rep list, sorted

The bundled file is backed up (timestamp-suffixed) before being overwritten.
"""

import os
import shutil
import sys

import openpyxl

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
BUNDLED_PATH = os.path.join(_APP_DIR, "LPM Salesforce and Overlay Collection Report.xlsx")

REP_LIST_SHEET_CANDIDATES = ["LPM Salesforce Collection Rep L", "LPM Salesforce Collection Rep List"]
OVERLAY_SHEET_CANDIDATES = ["Overlay Collection Report", "Overlay Report"]

REP_LIST_HEADERS = [
    "Site State", "Salesforce Collection Name", "Salesforce Collection ID",
    "Salesforce Collection Sales Person Name", "Salesforce Collection Salesperson ID",
    "Salesforce Collection Sales Person Employee Nbr", "Salesforce Collection Sales Person Status",
    "Salesforce Collection Team Nbr", "Salesforce Collection Team Name",
    "Salesforce Collection Sales Person Territory",
]
OVERLAY_HEADERS = [
    "Site State", "Overlay Assignment Name", "Overlay Assignment Type",
    "Overlay Assignment Salesperson Type Desc", "Overlay Assignment MMR Salesperson Name",
    "Overlay Assignment MMR Salesperson ID",
]
COLLECTION_ID_HEADERS = ["Site State", "Salesforce Collection Name", "Salesforce Collection ID"]


def find_sheet(wb, candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


def find_header_row(ws, expected_first_header):
    """Return (header_row_idx, col_offset) — col_offset skips a leading blank column."""
    for r in range(1, 6):
        for offset in (0, 1):
            val = ws.cell(r, 1 + offset).value
            if val and str(val).strip() == expected_first_header:
                return r, offset
    raise ValueError(f"Could not locate header row starting with {expected_first_header!r}")


def extract_rows(ws, n_cols, expected_first_header):
    header_row, offset = find_header_row(ws, expected_first_header)
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(r, c + offset).value for c in range(1, n_cols + 1)]
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(row)
    return rows


def write_sheet(ws, headers, rows):
    ws.delete_rows(1, ws.max_row)
    ws.append([None] * len(headers))
    ws.append(headers)
    for row in rows:
        ws.append(row)


def main():
    if len(sys.argv) < 2:
        print("Usage: python update_collection_report.py <rep_list_export.xlsx> [overlay_export.xlsx]", file=sys.stderr)
        sys.exit(1)

    rep_list_path = sys.argv[1]
    overlay_path = sys.argv[2] if len(sys.argv) >= 3 else rep_list_path

    if not os.path.isfile(rep_list_path):
        print(f"ERROR: File not found: {rep_list_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading rep list export: {rep_list_path}")
    src_wb = openpyxl.load_workbook(rep_list_path, data_only=True)
    rep_ws = find_sheet(src_wb, REP_LIST_SHEET_CANDIDATES)
    if rep_ws is None:
        print(f"ERROR: No rep list sheet found. Sheets present: {src_wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    rep_rows = extract_rows(rep_ws, len(REP_LIST_HEADERS), "Site State")
    print(f"  Rep list rows: {len(rep_rows)}")

    overlay_wb = src_wb
    if overlay_path != rep_list_path:
        print(f"Reading overlay export: {overlay_path}")
        overlay_wb = openpyxl.load_workbook(overlay_path, data_only=True)
    overlay_ws = find_sheet(overlay_wb, OVERLAY_SHEET_CANDIDATES)
    overlay_rows = []
    if overlay_ws is not None:
        overlay_rows = extract_rows(overlay_ws, len(OVERLAY_HEADERS), "Site State")
    print(f"  Overlay rows: {len(overlay_rows)}")

    # Dedupe (state, name, id) for the Collection ID List, preserve string IDs.
    seen = set()
    collection_id_rows = []
    for row in rep_rows:
        state, name, cid = row[0], row[1], row[2]
        if not state or not name or not cid:
            continue
        key = (str(state).strip().upper(), str(name).strip(), str(cid).strip())
        if key in seen:
            continue
        seen.add(key)
        collection_id_rows.append(list(key))
    collection_id_rows.sort(key=lambda r: (r[0], r[1]))
    print(f"  Unique collection IDs: {len(collection_id_rows)}")

    if not os.path.isfile(BUNDLED_PATH):
        print(f"ERROR: Bundled report not found at {BUNDLED_PATH}", file=sys.stderr)
        sys.exit(1)

    backup_path = BUNDLED_PATH.replace(".xlsx", "_backup.xlsx")
    shutil.copy2(BUNDLED_PATH, backup_path)
    print(f"Backed up existing report to: {backup_path}")

    out_wb = openpyxl.load_workbook(BUNDLED_PATH)

    rep_out = find_sheet(out_wb, REP_LIST_SHEET_CANDIDATES)
    write_sheet(rep_out, REP_LIST_HEADERS, rep_rows)

    overlay_out = find_sheet(out_wb, OVERLAY_SHEET_CANDIDATES)
    write_sheet(overlay_out, OVERLAY_HEADERS, overlay_rows)

    coll_out = out_wb["Collection ID List"]
    write_sheet(coll_out, COLLECTION_ID_HEADERS, collection_id_rows)

    # Force Collection ID columns to text so precision isn't lost on reopen.
    for ws, id_col in ((rep_out, 3), (coll_out, 3)):
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(r, id_col)
            if cell.value is not None:
                cell.number_format = "@"

    out_wb.save(BUNDLED_PATH)
    print(f"Saved updated report: {BUNDLED_PATH}")
    print(f"  {rep_out.title}: {len(rep_rows)} rows")
    print(f"  {overlay_out.title}: {len(overlay_rows)} rows")
    print(f"  {coll_out.title}: {len(collection_id_rows)} rows")


if __name__ == "__main__":
    main()
