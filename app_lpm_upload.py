"""
Streamlit web app for the SPP LPM Upload Generator.

Run with:
    streamlit run app_lpm_upload.py
"""

import io
import os
import sys
import tempfile
from pathlib import Path

import openpyxl
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import generate_lpm_upload as gen

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="SPP LPM Upload Generator",
    page_icon="📊",
    layout="centered",
)

_APP_DIR = Path(__file__).resolve().parent
BUNDLED_COLLECTION = str(_APP_DIR / "LPM Salesforce and Overlay Collection Report.xlsx")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_goal_groups(xlsm_bytes):
    """Return unique non-blank goal groups from Tracking Table, in order."""
    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as f:
        f.write(xlsm_bytes)
        tmp = f.name
    try:
        wb = openpyxl.load_workbook(tmp, keep_vba=True, data_only=True, read_only=True)
        if "Tracking Table" not in wb.sheetnames:
            return []
        ws = wb["Tracking Table"]
        groups, seen = [], set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            g = str(row[0]).strip() if row[0] else ""
            if g and g not in seen:
                seen.add(g)
                groups.append(g)
        wb.close()
        return groups
    finally:
        os.unlink(tmp)


def get_state_collections(state, source):
    """
    Return {display_name: collection_id} for the given state.
    Delegates to gen._parse_collection_sheet_raw which reads raw XML to avoid
    openpyxl float64 precision loss. Returns original-cased display names.
    source: file path string or bytes.
    """
    try:
        import zipfile, xml.etree.ElementTree as ET
        NS     = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        PKG_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

        if isinstance(source, (bytes, bytearray)):
            zf = zipfile.ZipFile(io.BytesIO(source))
        else:
            zf = zipfile.ZipFile(str(source))

        # Map rId -> sheet name
        with zf.open("xl/workbook.xml") as f:
            wb_tree = ET.parse(f)
        rid_to_name = {}
        for sh in wb_tree.findall(f".//{{{NS}}}sheet"):
            rid = sh.get(f"{{{REL_NS}}}id")
            rid_to_name[rid] = sh.get("name")

        # Map rId -> file path
        rid_to_path = {}
        if "xl/_rels/workbook.xml.rels" in zf.namelist():
            with zf.open("xl/_rels/workbook.xml.rels") as f:
                rels_tree = ET.parse(f)
            for rel in rels_tree.findall(f"{{{PKG_NS}}}Relationship"):
                target = rel.get("Target")
                path = target.lstrip("/") if target.startswith("/") else "xl/" + target
                rid_to_path[rel.get("Id")] = path

        target = "Collection ID List"
        candidate = None
        for rid, name in rid_to_name.items():
            if name == target:
                candidate = rid_to_path.get(rid)
                break

        if not candidate or candidate not in zf.namelist():
            zf.close()
            return {}

        shared_strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            with zf.open("xl/sharedStrings.xml") as f:
                ss_tree = ET.parse(f)
            for si in ss_tree.findall(f"{{{NS}}}si"):
                parts = si.findall(f".//{{{NS}}}t")
                shared_strings.append("".join(p.text or "" for p in parts))

        def cell_val(c_elem):
            t = c_elem.get("t", "n")
            if t == "inlineStr":
                is_elem = c_elem.find(f"{{{NS}}}is")
                if is_elem is None:
                    return ""
                parts = is_elem.findall(f".//{{{NS}}}t")
                return "".join(p.text or "" for p in parts)
            v = c_elem.find(f"{{{NS}}}v")
            if v is None or v.text is None:
                return ""
            if t == "s":
                idx = int(v.text)
                return shared_strings[idx] if idx < len(shared_strings) else ""
            val = v.text.strip()
            if val.endswith(".0"):
                val = val[:-2]
            return val

        result = {}
        with zf.open(candidate) as f:
            ws_tree = ET.parse(f)
        for row_elem in ws_tree.findall(f".//{{{NS}}}row"):
            if int(row_elem.get("r", 0)) < 2:
                continue
            col_vals = {}
            for c in row_elem.findall(f"{{{NS}}}c"):
                ref = c.get("r", "")
                col = "".join(ch for ch in ref if ch.isalpha()).upper()
                col_vals[col] = cell_val(c)
            row_state = col_vals.get("A", "").strip().upper()
            name      = col_vals.get("B", "").strip()
            cid       = col_vals.get("C", "").strip()
            if row_state != state.upper():
                continue
            if not name or not cid:
                continue
            if "(do not use)" in name.lower():
                continue
            result[name] = cid
        zf.close()
        return result
    except Exception:
        return {}


def auto_match(group_name, state, collections):
    """
    Pre-select a collection for a goal group.
    Tries group_name as-is first (some Goal Builders already store the full
    Salesforce collection name, e.g. 'CI - SPP - MN - ALD On Premise'), then
    falls back to the CI prefix patterns 'CI - {STATE} SPP - {group}',
    'CI - {STATE} - SPP - {group}', and 'CI - SPP - {STATE} - {group}'
    (state after SPP instead of before) for Goal Builders that only store
    the short group label, e.g. 'Combo'.
    Falls back to None if no match found.
    """
    candidates = [
        group_name,
        f"CI - {state.upper()} SPP - {group_name}",
        f"CI - {state.upper()} - SPP - {group_name}",
        f"CI - SPP - {state.upper()} - {group_name}",
    ]
    for candidate in candidates:
        for cname in collections:
            if cname.lower() == candidate.lower():
                return cname
    return None


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("How to use")
    st.markdown("""
1. Upload your **Goal Builder** `.xlsm` file
2. Confirm the **Collection ID Mapping** for each Goal Group
3. Click **Generate CSV**
4. Download the output and review any skipped rows

---
**State** is read from the first word of the Goal Builder filename
(e.g. `SD SPP Goal Builder.xlsm` → **SD**)

The **LPM Collection Report** is bundled automatically.
Upload a replacement below only if you have a newer version.
""")
    st.markdown("---")
    st.caption("LPM Upload Generator · Southern Glazer's")

# ---------------------------------------------------------------------------
# File uploads
# ---------------------------------------------------------------------------
st.title("SPP LPM Upload Generator")

goal_builder_file = st.file_uploader(
    "Goal Builder (.xlsm)",
    type=["xlsm"],
    help="State abbreviation is derived from the filename.",
)

collection_file = st.file_uploader(
    "Override Collection Report (.xlsx) — optional",
    type=["xlsx"],
    help="Leave blank to use the bundled collection report.",
)

# Reset session state when a new Goal Builder is uploaded
if "last_gb_name" not in st.session_state:
    st.session_state.last_gb_name = None

if goal_builder_file and goal_builder_file.name != st.session_state.last_gb_name:
    st.session_state.last_gb_name   = goal_builder_file.name
    st.session_state.goal_groups    = None
    st.session_state.manual_mapping = {}
    st.session_state.result         = None

# ---------------------------------------------------------------------------
# Collection ID Mapping dropdowns
# ---------------------------------------------------------------------------
if goal_builder_file:
    state = gen.derive_state_from_filename(goal_builder_file.name)

    # Determine collection source
    if collection_file:
        coll_source = collection_file.getvalue()
        coll_source_label = f"Uploaded: {collection_file.name}"
    elif Path(BUNDLED_COLLECTION).is_file():
        coll_source = BUNDLED_COLLECTION
        coll_source_label = f"Bundled: {Path(BUNDLED_COLLECTION).name}"
    else:
        coll_source = None
        coll_source_label = "No collection report found"

    collections = get_state_collections(state, coll_source) if coll_source else {}
    st.caption(f"Collection report: {coll_source_label}")

    # Parse goal groups once and cache in session state
    if st.session_state.get("goal_groups") is None:
        with st.spinner("Reading Goal Builder…"):
            st.session_state.goal_groups = extract_goal_groups(goal_builder_file.getvalue())

    goal_groups = st.session_state.goal_groups

    if goal_groups and collections:
        with st.expander("Collection ID Mapping", expanded=True):
            st.caption(
                f"State: **{state}** — {len(collections)} collection(s) available. "
                "Match each Goal Group to its Salesforce Collection ID."
            )
            options = ["(none)"] + list(collections.keys())
            manual_mapping = {}

            st.markdown(
                """
                <style>
                [data-testid="stVerticalBlockBorderWrapper"] .collection-scroll {
                    max-height: 400px;
                    overflow-y: auto;
                    padding-right: 8px;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )
            with st.container(height=400):
                for group in goal_groups:
                    best = auto_match(group, state, collections)
                    default_idx = options.index(best) if best and best in options else 0
                    selected = st.selectbox(
                        group,
                        options,
                        index=default_idx,
                        key=f"cmap_{group}",
                    )
                    if selected != "(none)":
                        manual_mapping[group.lower()] = collections[selected]

            st.session_state.manual_mapping = manual_mapping

    elif goal_groups and not collections:
        st.info(
            f"State: **{state}** — no collections found in the collection report. "
            "`salesforce_collection_ids` will be blank."
        )

# ---------------------------------------------------------------------------
# Generate button
# ---------------------------------------------------------------------------
generate_clicked = st.button(
    "Generate CSV",
    type="primary",
    disabled=(goal_builder_file is None),
    use_container_width=True,
)

# ---------------------------------------------------------------------------
# Processing
# ---------------------------------------------------------------------------
if generate_clicked and goal_builder_file is not None:
    with st.spinner("Processing Tracking Table…"):
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                gb_path = os.path.join(tmpdir, goal_builder_file.name)
                with open(gb_path, "wb") as f:
                    f.write(goal_builder_file.getvalue())

                output_path  = os.path.join(tmpdir, "lpm_upload.csv")
                skipped_path = os.path.join(tmpdir, "lpm_skipped.csv")

                # Apply the manual mapping from the dropdowns
                state = gen.derive_state_from_filename(gb_path)
                gen.COLLECTION_LOOKUP = st.session_state.get("manual_mapping", {})

                records, skipped, header_row = gen.load_tracking_table(gb_path)
                if not records:
                    raise ValueError("No valid records found in the Tracking Table after filtering.")

                order, groups = gen.group_records(records)
                total_trackers, total_ptgs = gen.generate_output(order, groups, output_path)

                if skipped:
                    gen.write_skipped_csv(skipped_path, skipped, header_row)

                with open(output_path, "rb") as f:
                    output_bytes = f.read()

                skipped_bytes = None
                if skipped and os.path.exists(skipped_path):
                    with open(skipped_path, "rb") as f:
                        skipped_bytes = f.read()

                collection_info = [
                    f"{group}  →  {cid}"
                    for group, cid in sorted(gen.COLLECTION_LOOKUP.items())
                ]

            st.session_state.result = {
                "error":           None,
                "state":           state,
                "total_trackers":  total_trackers,
                "total_ptgs":      total_ptgs,
                "skipped":         skipped,
                "output_bytes":    output_bytes,
                "skipped_bytes":   skipped_bytes,
                "base_name":       os.path.splitext(goal_builder_file.name)[0],
                "collection_info": collection_info,
            }

        except Exception as exc:
            st.session_state.result = {"error": str(exc)}

# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------
result = st.session_state.get("result")

if result:
    st.markdown("---")

    if result.get("error"):
        st.error(f"**Error:** {result['error']}")

    else:
        if result["collection_info"]:
            with st.expander(
                f"State: **{result['state']}** — {len(result['collection_info'])} collection ID(s) applied",
                expanded=False,
            ):
                for line in result["collection_info"]:
                    st.text(line)
        else:
            st.info("`salesforce_collection_ids` will be blank — no collections were mapped.")

        c1, c2, c3 = st.columns(3)
        c1.metric("Trackers", result["total_trackers"])
        c2.metric("PTGs", result["total_ptgs"])
        c3.metric("Skipped rows", len(result["skipped"]))

        if result["skipped"]:
            st.warning(f"⚠️ {len(result['skipped'])} row(s) were skipped — review before uploading.")
            with st.expander("View skipped rows"):
                for entry in result["skipped"]:
                    goal_group = (entry["raw_row"][0] or "") if entry["raw_row"] else ""
                    st.markdown(
                        f"**Row {entry['row_num']}** &nbsp;·&nbsp; "
                        f"`{goal_group}` &nbsp;·&nbsp; {entry['reason']}"
                    )
        else:
            st.success("✅ All rows processed — no skipped rows.")

        st.markdown("")

        base = result["base_name"]
        st.download_button(
            label="⬇️  Download LPM Upload CSV",
            data=result["output_bytes"],
            file_name=f"{base}_lpm_upload.csv",
            mime="text/csv",
            use_container_width=True,
        )
        if result["skipped_bytes"]:
            st.download_button(
                label="⬇️  Download Skipped Rows CSV",
                data=result["skipped_bytes"],
                file_name=f"{base}_skipped.csv",
                mime="text/csv",
                use_container_width=True,
            )
